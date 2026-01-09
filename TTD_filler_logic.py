#!/usr/bin/env python3
"""
TTD TEMPLATE FILLER – FINAL STABLE VERSION

✔ Address line 3: removes city/state/pincode
✔ Address line 2 fallback → Receiver City
✔ L/B/H mapped from volumetric file by Category + Quantity
✔ 6 Sheet Calendar volumetric logic added
✔ Barcode preserved per postal row
✔ Receiver City / State untouched
✔ CLI arguments unchanged
"""

import argparse
import pandas as pd
from openpyxl import load_workbook
import re

# --------------------------------------------------
# HELPERS
# --------------------------------------------------

def clean_mobile(mobile):
    if pd.isna(mobile):
        return ""
    digits = re.sub(r"\D", "", str(mobile))
    if len(digits) > 10 and digits.startswith("91"):
        digits = digits[2:]
    return digits if len(digits) == 10 else ""

def split_address(addr):
    if pd.isna(addr) or not addr:
        return "", "", ""

    parts = [p.strip() for p in str(addr).split(",") if p.strip()]

    # remove last 3: city, state, pincode
    if len(parts) > 3:
        parts = parts[:-3]

    line1 = parts[0] if len(parts) > 0 else ""
    line2 = parts[1] if len(parts) > 1 else ""
    line3 = ", ".join(parts[2:]) if len(parts) > 2 else ""

    return line1, line2, line3

# --------------------------------------------------
# VOLUMETRIC HANDLING
# --------------------------------------------------

def load_volumetric_tables(path):
    df = pd.read_excel(path, header=None)

    def block(r1, r2, c1, c2):
        b = df.iloc[r1:r2, c1:c2].copy()
        b.columns = ["Quantity", "L", "B", "H", "Weight"]
        b["Quantity"] = pd.to_numeric(b["Quantity"], errors="coerce")
        b = b.dropna(subset=["Quantity"])
        b["Quantity"] = b["Quantity"].astype(int)
        return b.set_index("Quantity")

    return {
        "calendar": block(2, 22, 0, 5),
        "ttc": block(2, 27, 6, 11),
        "big_diary": block(28, 48, 0, 5),
        "small_diary": block(28, 53, 6, 11),
    }

def get_dimensions(vol, category, qty):
    qty = int(qty)
    c = str(category).lower()

    # 🔹 NEW: 6 Sheet Calendar
    if "6 sheet" in c:
        base = vol["calendar"].iloc[0]
        L = int(base["L"])
        B = int(base["B"])

        packs = min(qty, 10)
        H = int(2 + (packs - 1) * 0.5)

        return L, B, H

    if "calendar" in c and "table" not in c:
        table = vol["calendar"]
    elif "table" in c:
        table = vol["ttc"]
    elif "big" in c:
        table = vol["big_diary"]
    elif "small" in c:
        table = vol["small_diary"]
    else:
        return None, None, None

    if qty in table.index:
        r = table.loc[qty]
    else:
        r = table[table.index <= qty].iloc[-1]

    return int(r["L"]), int(r["B"]), int(r["H"])

# --------------------------------------------------
# MAIN
# --------------------------------------------------

def main(args):

    # -------- POSTAL --------
    postal = pd.read_excel(args.input2, header=3)

    tr_col      = postal.columns[1]
    name_col    = postal.columns[2]
    addr_col    = postal.columns[3]
    city_col    = postal.columns[4]
    pin_col     = postal.columns[5]
    mobile_col  = postal.columns[6]
    qty_col     = postal.columns[7]
    weight_col  = postal.columns[8]
    barcode_col = postal.columns[9]

    postal = postal[
        [tr_col, name_col, addr_col, city_col, pin_col,
         mobile_col, qty_col, weight_col, barcode_col]
    ].copy()

    postal["__TR"] = postal[tr_col].astype(str).str.strip()
    postal["Receiver name"] = postal[name_col]
    postal["Full Address"] = postal[addr_col]
    postal["Receiver city"] = postal[city_col]
    postal["Receiver pincode"] = pd.to_numeric(
        postal[pin_col], errors="coerce"
    ).fillna(0).astype(int)
    postal["Receiver mobile"] = postal[mobile_col].apply(clean_mobile)
    postal["Quantity"] = pd.to_numeric(
        postal[qty_col], errors="coerce"
    ).fillna(1).astype(int)
    postal["Physical weight in grams"] = pd.to_numeric(
        postal[weight_col], errors="coerce"
    ).fillna(0).astype(int)
    postal["Barcode"] = postal[barcode_col].astype(str).str.strip()

    merged = postal[
        postal["Receiver pincode"].between(100000, 999999)
    ].copy()

    # -------- ORDERS --------
    orders = pd.read_excel(args.input1, sheet_name="Publications_Report")
    orders["__TR"] = orders["Booking No"].astype(str).str.strip()
    orders = orders[["__TR", "State", "Category"]].drop_duplicates("__TR")

    merged = merged.merge(
        orders, on="__TR", how="left", validate="many_to_one"
    )
    merged["State"] = merged["State"].fillna("Tamil Nadu")

    # -------- VOLUMETRIC --------
    vol = load_volumetric_tables(args.volumetric)

    # -------- TEMPLATE --------
    wb = load_workbook(args.template)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    defaults = [c.value for c in ws[2]]

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in r:
            c.value = None

    serial_no = 1

    for idx, row in merged.iterrows():
        excel_row = 2 + idx

        L, B, H = get_dimensions(
            vol, row["Category"], row["Quantity"]
        )

        a1, a2, a3 = split_address(row["Full Address"])

        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=excel_row, column=i)
            key = "" if h is None else h.lower()

            if "serial" in key:
                cell.value = serial_no
            elif "barcode" in key:
                cell.value = row["Barcode"]
            elif "physical weight" in key:
                cell.value = row["Physical weight in grams"]
            elif "receiver city" in key:
                cell.value = row["Receiver city"]
            elif "receiver pincode" in key:
                cell.value = row["Receiver pincode"]
            elif "receiver name" in key:
                cell.value = row["Receiver name"]
            elif "receiver mobile" in key:
                cell.value = row["Receiver mobile"]
            elif "add line 1" in key:
                cell.value = a1
            elif "add line 2" in key:
                cell.value = a2 if a2 else row["Receiver city"]
            elif "add line 3" in key:
                cell.value = a3
            elif "length" in key:
                cell.value = L
            elif "breadth" in key or "diameter" in key:
                cell.value = B
            elif "height" in key:
                cell.value = H
            elif "receiver state" in key:
                cell.value = row["State"]
            else:
                cell.value = defaults[i - 1]

            if "sender mobile" in key:
                cell.value = 1234567890
            if "sender state" in key:
                cell.value = "Andhra Pradesh"
            if "sender add line 1" in key:
                cell.value = "SALES WING OF PUBLICATIONS"
            if "sender add line 2" in key:
                cell.value = "TTD PRESS COMPOUND"
            if "sender add line 3" in key:
                cell.value = "Tirupati-517507"

        serial_no += 1

    wb.save(args.output)
    print(f"FINAL PERFECT OUTPUT GENERATED: {args.output}")
    print(f"TOTAL ARTICLES GENERATED : {serial_no - 1}")

# --------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input1", required=True)
    parser.add_argument("--input2", required=True)
    parser.add_argument("--template", required=True)
    parser.add_argument("--volumetric", required=True)
    parser.add_argument("--output", required=True)
    main(parser.parse_args())
